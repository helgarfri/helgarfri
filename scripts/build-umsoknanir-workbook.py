#!/usr/bin/env python3
"""Generate jobs/umsoknanir-2026.xlsx; umsóknaryfirlit með staða-lista."""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

LINK_STYLE = Font(color="0563C1", underline="single")

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "jobs" / "umsoknanir-2026.xlsx"

HEADERS = ["Fyrirtæki", "Starfstitill", "Umsóknarfrestur", "Linkur", "Staða"]

ROWS: list[tuple[str, str, str, str, str]] = [
    (
        "Advania",
        "Sumarstörf hjá Advania",
        "15.04.2026",
        "",
        "Sótt um",
    ),
    (
        "Origo",
        "Almenn umsókn",
        "31.12.2026",
        "",
        "Ekki sótt um",
    ),
    (
        "atNorth",
        "DevOps & Systems Engineer",
        "",
        "",
        "Ekki sótt um",
    ),
    (
        "Asana",
        "Software Engineer, Infrastructure (Reykjavík)",
        "",
        "https://asana.com/jobs/apply/5480143",
        "Ekki sótt um",
    ),
    (
        "Infomentor",
        "Full-stack forritari í menntatækni",
        "23.03.2026",
        "https://alfred.is/starf/full-stack-forritari-i-menntataekni",
        "Ekki sótt um",
    ),
    (
        "Chargitect",
        "Software Engineer (Bergen)",
        "",
        "https://arbeidsplassen.nav.no/stillinger/stilling/30f23dbd-73cf-413a-9643-c0e820429f5e",
        "Ekki sótt um",
    ),
    (
        "Monil AS",
        "Frontend Software Engineer (React Native, web)",
        "",
        "https://careers.monil.com/jobs/6457977-frontend-software-engineer-react-native-web",
        "Ekki sótt um",
    ),
    (
        "Visoid",
        "Software Engineer (Oslo)",
        "",
        "https://visoid.homerun.co/software-engineer-3/en",
        "Ekki sótt um",
    ),
    (
        "Coody",
        "Fullstack Developer (TypeScript), Stockholm",
        "",
        "https://www.coody.com/careers",
        "Ekki sótt um",
    ),
    (
        "Yazen",
        "Full Stack Engineer (remote)",
        "",
        "https://thehub.io/jobs/69127cd4cc00a86a1005c2d6",
        "Ekki sótt um",
    ),
    (
        "Mercive",
        "Full Stack Developer (Copenhagen)",
        "",
        "https://thehub.io/jobs/69b90d488b60f3e05fb92945",
        "Ekki sótt um",
    ),
    (
        "Shine",
        "Full Stack Software Engineer, Banking Experience",
        "",
        "https://thehub.io/jobs/6993b01d94862e1d27011b23",
        "Ekki sótt um",
    ),
    (
        "Veo Technologies",
        "Full Stack Engineer, Monetization (Copenhagen)",
        "",
        "https://thehub.io/jobs/68e8538b6b6a942af7bb69d3",
        "Ekki sótt um",
    ),
    (
        "Gildi",
        "Forritari",
        "07.04.2026",
        "https://recruitcrm.io/apply/17740169183520016834FnD?utm_source=alfred_jobs&utm_medium=job_platform&utm_campaign=forritari-50&utm_content=164256",
        "Ekki sótt um",
    ),
    (
        "Rapyd Europe",
        "NOC Specialist (sumar)",
        "þar til ráðið",
        "https://alfred.is/starf/noc-specialist-summer-position",
        "Ekki sótt um",
    ),
    (
        "Total Tempo ehf.",
        "Full stack forritari",
        "",
        "https://alfred.is/starf/full-stack-forritari-4",
        "Ekki sótt um",
    ),
    (
        "The Flex",
        "Full Stack Engineer (Web & APIs)",
        "",
        "https://jobs.ashbyhq.com/The-Flex/79cf4d28-e152-4817-b690-7d2ad5992076",
        "Ekki sótt um",
    ),
]

# Aukalínur fyrir ný fyrirtæki
EMPTY_TEMPLATE_ROWS = 8

STATUS_CHOICES = [
    "Ekki sótt um",
    "Í vinnslu",
    "Sótt um",
    "Viðtal",
    "Samþykkt",
    "Hafnað",
    "Afturkallað",
    "Engin svörun",
]


def main() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Umsóknir"

    ws.append(HEADERS)
    for row in ROWS:
        ws.append(list(row))
    for _ in range(EMPTY_TEMPLATE_ROWS):
        ws.append(["", "", "", "", "Ekki sótt um"])

    last_row = ws.max_row

    header_fill = PatternFill("solid", fgColor="F2F2F2")
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=len(HEADERS)):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for r in range(2, last_row + 1):
        link_cell = ws.cell(row=r, column=4)
        url = (link_cell.value or "").strip()
        if url.startswith("http://") or url.startswith("https://"):
            link_cell.hyperlink = url
            link_cell.value = url
            link_cell.font = LINK_STYLE

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 56
    ws.column_dimensions["E"].width = 20

    # Listaval í dálki E (virkar í Excel / LibreOffice Calc með gagnavalideringu)
    quoted = ",".join(STATUS_CHOICES)
    dv = DataValidation(
        type="list",
        formula1=f'"{quoted}"',
        allow_blank=True,
    )
    dv.error = "Veldu gildi úr listanum eða skildu reitinn eftir auðan."
    dv.errorTitle = "Ógilt val"
    ws.add_data_validation(dv)
    dv.add(f"E2:E{last_row}")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{last_row}"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
